from __future__ import annotations

import json
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Set, Tuple

import numpy as np
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config
from db import get_recent_snapshots, insert_snapshot, update_session_report_path as update_session_raport_path


def _border_subtire() -> Border:
    side = Side(style="thin", color="A6A6A6")
    return Border(left=side, right=side, top=side, bottom=side)


def _stil_titlu(cell, fill: str = "6F5B95") -> None:
    cell.font = Font(color="FFFFFF", bold=True, size=12)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _border_subtire()


def _stil_antet(cell, fill: str = "20B7D7") -> None:
    cell.font = Font(color="FFFFFF", bold=True)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _border_subtire()


def _stil_nota(cell, fill: str = "FFF2CC") -> None:
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(color="000000", italic=True)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = _border_subtire()


def _latime_automata(ws, min_width: int = 12, max_width: int = 40) -> None:
    for col_cells in ws.columns:
        first = col_cells[0]
        if getattr(first, "column", None) is None:
            continue
        col_letter = get_column_letter(first.column)
        max_len = 0
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def _fill_stare(stare: str) -> PatternFill:
    stare = (stare or "").lower()
    if stare == "critical":
        return PatternFill("solid", fgColor="F4CCCC")
    if stare == "warning":
        return PatternFill("solid", fgColor="FCE5CD")
    if stare in {"fatigue risk", "attention drop"}:
        return PatternFill("solid", fgColor="FFF2CC")
    return PatternFill("solid", fgColor="D9EAD3")


def _rotunjeste(value: Any, digits: int = 2) -> Any:
    try:
        return round(float(value), digits)
    except Exception:
        return value


def _format_session_datetime(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(value, (int, float)):
        try:
            return datetime.fromtimestamp(float(value)).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return str(value)

    text = str(value).strip()
    if not text:
        return ""

    if text.replace(".", "", 1).isdigit():
        try:
            return datetime.fromtimestamp(float(text)).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            pass

    normalized = text.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(normalized).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return text


def _student_id_excel_value(value: Any) -> Any:
    text = str(value).strip()
    if text.isdigit():
        try:
            return int(text)
        except Exception:
            return text
    return text


def _incarca_metrici_evaluare() -> Dict[str, Any]:
    metrics_path = Path(
        getattr(
            config,
            "EVAL_METRICS_PATH",
            Path(getattr(config, "SESSION_REPORTS_DIR", "reports")) / "last_eval_metrics.json",
        )
    )

    if not metrics_path.exists():
        return {}

    try:
        with open(metrics_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _bool_option(value: Any, default: bool = False) -> bool:
    if value is None:
        return bool(default)
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)

    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "on", "checked"}:
        return True
    if text in {"0", "false", "no", "n", "off", "unchecked", ""}:
        return False
    return bool(default)


def _normalize_report_options(report_options: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    opts = dict(report_options or {})

    report_type = str(
        opts.get("report_type", opts.get("reportType", "developer")) or "developer"
    ).lower()
    if report_type in {"user", "prof"}:
        report_type = "teacher"
    if report_type not in {"teacher", "developer"}:
        report_type = "teacher"
    opts["report_type"] = report_type

    charts_raw = opts.get("charts", {})
    charts = dict(charts_raw) if isinstance(charts_raw, dict) else {}

    if report_type == "developer":
        opts["charts"] = {"individual": True, "group": True, "alerts": True}
        return opts

    individual_default = True
    group_default = True
    alerts_default = False

    opts["charts"] = {
        "individual": _bool_option(
            charts.get(
                "individual",
                opts.get("chart_individual", opts.get("chartsIndividual", individual_default)),
            ),
            individual_default,
        ),
        "group": _bool_option(
            charts.get(
                "group",
                opts.get("chart_group", opts.get("chartsGroup", group_default)),
            ),
            group_default,
        ),
        "alerts": _bool_option(
            charts.get(
                "alerts",
                opts.get("chart_alerts", opts.get("chartsAlerts", alerts_default)),
            ),
            alerts_default,
        ),
    }
    return opts



class RaportManager:
    def __init__(
        self,
        db_path: str,
        recent_rapoarte_limit: int = 3,
        interval_raport_s: int = 60,
    ) -> None:
        self.db_path = db_path
        self.recent_rapoarte_limit = int(recent_rapoarte_limit)
        self.interval_raport_s = int(interval_raport_s)
        self.reset_runtime()

    def reset_runtime(self) -> None:
        self.raport_buffer: List[Dict[str, Any]] = []
        self.trend_buffer: List[Dict[str, Any]] = []
        self.last_raport_ts: float = time.time()
        self.recent_rapoarte: List[Dict[str, Any]] = []

    def bind_session(self, session_id: Optional[int]) -> None:
        self.reset_runtime()
        if session_id is None:
            return
        self.recent_rapoarte = get_recent_snapshots(
            self.db_path,
            session_id=session_id,
            limit=self.recent_rapoarte_limit,
        )

    def adauga_esantion_runtime(
        self,
        now: float,
        faces: int,
        heads: int,
        fatigue: float,
        attention: float,
        fatigue_alert_active: bool,
        attention_alert_active: bool,
        student_alerts: List[dict],
        studenti_runtime: List[dict],
        class_decision_explanation: str,
    ) -> None:
        self.raport_buffer.append(
            {
                "ts": float(now),
                "faces": float(faces),
                "heads": float(heads),
                "fatigue": float(fatigue),
                "attention": float(attention),
                "fatigue_alert_active": bool(fatigue_alert_active),
                "attention_alert_active": bool(attention_alert_active),
                "student_alert_count": int(len(student_alerts or [])),
                "critical_student_count": int(
                    sum(1 for x in (student_alerts or []) if str(x.get("severity", "")) == "critical")
                ),
                "decision_explanation": str(class_decision_explanation or ""),
            }
        )

        puncte_studenti: List[Dict[str, Any]] = []
        vazuti: Set[str] = set()
        for item in studenti_runtime or []:
            student_id = str(item.get("student_id", "")).strip()
            if not student_id or student_id in vazuti:
                continue
            vazuti.add(student_id)
            puncte_studenti.append(
                {
                    "student_id": student_id,
                    "fatigue_pct": _rotunjeste(item.get("fatigue_pct", 0.0)),
                    "attention_pct": _rotunjeste(item.get("attention_pct", 0.0)),
                }
            )

        self.trend_buffer.append(
            {
                "ts": float(now),
                "timestamp_label": datetime.fromtimestamp(float(now)).strftime("%H:%M:%S"),
                "group_fatigue_pct": _rotunjeste(fatigue),
                "group_attention_pct": _rotunjeste(attention),
                "group_people_count": int(heads if heads is not None else faces),
                "group_student_alert_count": int(len(student_alerts or [])),
                "studenti": puncte_studenti,
            }
        )

    def poate_stoca_snapshot(
        self,
        session_id: Optional[int],
        now: float,
        class_alert_type_and_level_fn: Callable[[bool, bool, int], Tuple[str, str]],
        class_reason_and_decision_fn: Callable[[float, float, bool, bool, int, str], Tuple[str, str]],
    ) -> None:
        if session_id is None:
            return

        interval_s = self.interval_raport_s if self.interval_raport_s > 0 else 60
        if now - self.last_raport_ts < interval_s:
            return

        if not self.raport_buffer:
            self.last_raport_ts = now
            return

        faces_avg = float(np.mean([x["faces"] for x in self.raport_buffer]))
        heads_avg = float(np.mean([x["heads"] for x in self.raport_buffer]))
        fatigue_avg = float(np.mean([x["fatigue"] for x in self.raport_buffer]))
        attention_avg = float(np.mean([x["attention"] for x in self.raport_buffer]))
        fatigue_max = float(np.max([x["fatigue"] for x in self.raport_buffer]))
        attention_min = float(np.min([x["attention"] for x in self.raport_buffer]))
        student_alert_count = int(max(x["student_alert_count"] for x in self.raport_buffer))
        critical_student_count = int(max(x["critical_student_count"] for x in self.raport_buffer))

        last = self.raport_buffer[-1]
        class_alert_type, class_alert_level = class_alert_type_and_level_fn(
            bool(last["fatigue_alert_active"]),
            bool(last["attention_alert_active"]),
            int(round(faces_avg)),
        )
        reason, decision = class_reason_and_decision_fn(
            fatigue_avg,
            attention_avg,
            bool(last["fatigue_alert_active"]),
            bool(last["attention_alert_active"]),
            int(round(faces_avg)),
            str(last.get("decision_explanation", "")),
        )

        insert_snapshot(
            db_path=self.db_path,
            session_id=session_id,
            created_at=now,
            faces_avg=faces_avg,
            heads_avg=heads_avg,
            active_students_avg=faces_avg,
            fatigue_avg=fatigue_avg,
            attention_avg=attention_avg,
            fatigue_max=fatigue_max,
            attention_min=attention_min,
            class_alert_type=class_alert_type,
            class_alert_level=class_alert_level,
            student_alert_count=student_alert_count,
            critical_student_count=critical_student_count,
            reason=reason,
            decision=decision,
        )

        self.recent_rapoarte = get_recent_snapshots(
            self.db_path,
            session_id=session_id,
            limit=self.recent_rapoarte_limit,
        )

        self.raport_buffer = []
        self.last_raport_ts = now

    def _stare_din_student(self, student: Dict[str, Any]) -> str:
        severity = str(student.get("final_severity", "none") or "none").lower()
        alert_type = str(student.get("final_alert_type", "none") or "none").lower()
        attention = float(student.get("attention_avg", 100.0) or 100.0)
        fatigue = float(student.get("fatigue_avg", 0.0) or 0.0)

        if attention < 30 and fatigue > 60:
            return "Critical"
        if severity == "critical" or severity == "warning":
            return "Warning"
        if alert_type == "fatigue":
            return "Fatigue risk"
        if alert_type == "attention":
            return "Attention drop"
        return "Monitoring"

    def _actiune_recomandata(self, student: Dict[str, Any]) -> str:
        severity = str(student.get("final_severity", "none") or "none").lower()
        alert_type = str(student.get("final_alert_type", "none") or "none").lower()

        if severity == "critical":
            return "Individual follow-up is recommended"
        if severity == "warning":
            return "Closer monitoring is recommended"
        if alert_type == "fatigue":
            return "Monitor for fatigue-related signs"
        if alert_type == "attention":
            return "Apply attention refocusing strategies"
        return "Continue regular monitoring"

    def _humanize_reason(self, reason: str, attention: float, fatigue: float) -> str:
        text = str(reason or "").strip().lower()

        if attention <= 20:
            return "Student had major difficulty staying focused throughout the session."
        if attention <= 40:
            return "Student seemed distracted for a significant part of the session."
        if attention <= 60:
            return "Student showed fluctuating attention during the session."

        if fatigue >= 80:
            return "Student showed strong signs of fatigue and low energy."
        if fatigue >= 60:
            return "Student appeared tired during several moments of the session."
        if fatigue >= 40:
            return "Student showed mild signs of fatigue."

        if "critical" in text:
            return "Student struggled significantly to remain engaged."
        if "inattentive" in text:
            return "Student seemed distracted quite often during the session."
        if "fatigue" in text or "tired" in text or "sleepy" in text:
            return "Student showed visible signs of tiredness."
        if "attention" in text or "focus" in text:
            return "Student had some difficulty maintaining consistent attention."

        if text:
            cleaned = reason.strip()
            return cleaned[:1].upper() + cleaned[1:] if cleaned else "No major issues observed."

        return "No major issues observed."

    def _student_ids_valizi_pentru_raport(
        self,
        session_students: List[Dict[str, Any]],
    ) -> List[str]:
        min_samples = int(getattr(config, "MIN_SAMPLES_FOR_REPORT", 15))
        min_visible_ratio = float(getattr(config, "MIN_VISIBLE_RATIO_FOR_REPORT", 0.15))
        min_span_seconds = float(getattr(config, "MIN_SPAN_SECONDS_FOR_REPORT", 8.0))

        total_rows = len(self.trend_buffer)
        if total_rows <= 0:
            ids = []
            for student in session_students or []:
                sid = str(student.get("student_id", "")).strip()
                if sid:
                    ids.append(sid)
            return sorted(set(ids), key=lambda x: int(x) if x.isdigit() else x)

        per_student: Dict[str, Dict[str, Any]] = {}

        for punct in self.trend_buffer:
            ts = float(punct.get("ts", 0.0))
            for item in punct.get("studenti", []) or []:
                sid = str(item.get("student_id", "")).strip()
                if not sid:
                    continue

                entry = per_student.setdefault(
                    sid,
                    {
                        "samples": 0,
                        "first_ts": None,
                        "last_ts": None,
                    },
                )

                has_metric = (
                    item.get("fatigue_pct") is not None
                    or item.get("attention_pct") is not None
                )
                if not has_metric:
                    continue

                entry["samples"] += 1
                if entry["first_ts"] is None:
                    entry["first_ts"] = ts
                entry["last_ts"] = ts

        valid_ids: List[str] = []

        for sid, info in per_student.items():
            samples = int(info.get("samples", 0))
            visible_ratio = samples / total_rows if total_rows > 0 else 0.0

            first_ts = info.get("first_ts")
            last_ts = info.get("last_ts")
            span_seconds = 0.0
            if first_ts is not None and last_ts is not None:
                span_seconds = max(0.0, float(last_ts) - float(first_ts))

            if (
                samples >= min_samples
                and visible_ratio >= min_visible_ratio
                and span_seconds >= min_span_seconds
            ):
                valid_ids.append(sid)

        if not valid_ids and per_student:
            best_sid = max(
                per_student.items(),
                key=lambda kv: (
                    int(kv[1].get("samples", 0)),
                    float((kv[1].get("last_ts") or 0.0)) - float((kv[1].get("first_ts") or 0.0)),
                ),
            )[0]
            valid_ids = [best_sid]

        return sorted(set(valid_ids), key=lambda x: int(x) if x.isdigit() else x)

    def _filtreaza_studenti_sumar(
        self,
        session_students: List[Dict[str, Any]],
        valid_student_ids: List[str],
    ) -> List[Dict[str, Any]]:
        valid_set = set(valid_student_ids)
        if not valid_set:
            return []

        return [
            student
            for student in (session_students or [])
            if str(student.get("student_id", "")).strip() in valid_set
        ]

    def _construieste_sheet_overview(
        self,
        wb: Workbook,
        session_summary: Dict[str, Any],
        report_options: Optional[Dict[str, Any]] = None,
    ) -> None:
        report_options = report_options or {}
        report_type = str(report_options.get("report_type", "developer") or "developer").lower()
        teacher_mode = report_type in {"teacher", "prof", "user"}

        ws = wb.active
        ws.title = "Session Overview"
        ws.sheet_view.showGridLines = True
        ws.freeze_panes = "A3"

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 78

        ws.merge_cells("A1:B1")
        ws["A1"] = "Class Monitor - Session Overview"
        _stil_titlu(ws["A1"])

        eval_metrics = _incarca_metrici_evaluare()
        people_present = int(session_summary.get("unique_faces_seen", 0) or 0)
        max_heads = int(session_summary.get("max_heads_seen", 0) or 0)
        max_faces = int(session_summary.get("max_faces_seen", 0) or 0)
        max_people = max(max_heads, max_faces, people_present)

        if teacher_mode:
            rows = [
                ("Started at", _format_session_datetime(session_summary.get("started_at", ""))),
                ("Stopped at", _format_session_datetime(session_summary.get("stopped_at", ""))),
                ("Duration (minutes)", _rotunjeste(float(session_summary.get("duration_s", 0.0)) / 60.0, 2)),
                ("Estimated people present", people_present),
                ("Max people detected", max_people),
                ("Average fatigue (%)", _rotunjeste(session_summary.get("avg_fatigue", 0.0))),
                ("Average attention (%)", _rotunjeste(session_summary.get("avg_attention", 0.0))),
                ("Moments needing attention", int(session_summary.get("alert_event_count", session_summary.get("alert_count", 0)))),
            ]
        else:
            rows = [
                ("Started at", _format_session_datetime(session_summary.get("started_at", ""))),
                ("Stopped at", _format_session_datetime(session_summary.get("stopped_at", ""))),
                ("Duration (minutes)", _rotunjeste(float(session_summary.get("duration_s", 0.0)) / 60.0, 2)),
                ("Unique faces seen", people_present),
                ("Max faces seen", max_faces),
                ("Max heads seen", max_heads),
                ("Avg faces", _rotunjeste(session_summary.get("avg_faces", 0.0))),
                ("Avg heads", _rotunjeste(session_summary.get("avg_heads", 0.0))),
                ("Avg fatigue (%)", _rotunjeste(session_summary.get("avg_fatigue", 0.0))),
                ("Avg attention (%)", _rotunjeste(session_summary.get("avg_attention", 0.0))),
                ("Detection accuracy (%)", _rotunjeste(eval_metrics.get("accuracy_pct", ""))),
                ("Precision (%)", _rotunjeste(eval_metrics.get("precision_pct", ""))),
                ("Recall (%)", _rotunjeste(eval_metrics.get("recall_pct", ""))),
                ("F1 score (%)", _rotunjeste(eval_metrics.get("f1_pct", ""))),
                ("Moments needing attention", int(session_summary.get("alert_event_count", session_summary.get("alert_count", 0)))),
            ]

        start_row = 3
        for idx, (label, value) in enumerate(rows, start=start_row):
            ws.cell(idx, 1, label)
            ws.cell(idx, 2, value)
            ws.cell(idx, 1).font = Font(bold=True)
            ws.cell(idx, 1).fill = PatternFill("solid", fgColor="D9E2F3")
            ws.cell(idx, 1).border = _border_subtire()
            ws.cell(idx, 2).border = _border_subtire()
            ws.cell(idx, 2).alignment = Alignment(horizontal="right", vertical="center")

        _latime_automata(ws, min_width=14, max_width=42)

    def _construieste_sheet_studenti(
        self,
        wb: Workbook,
        session_students: List[Dict[str, Any]],
        valid_student_ids: List[str],
    ) -> None:
        ws = wb.create_sheet("Students Requiring Attention")
        ws.sheet_view.showGridLines = True
        ws.freeze_panes = "A3"

        ws.merge_cells("A1:H1")
        ws["A1"] = "Students Requiring Attention"
        _stil_titlu(ws["A1"])

        headers = [
            "Student ID",
            "Fatigue Avg",
            "Attention Avg",
            "Fatigue Max",
            "Attention Min",
            "Status",
            "Recommended Action",
            "Observation",
        ]

        for col_idx, header in enumerate(headers, start=1):
            _stil_antet(ws.cell(3, col_idx, header))

        rand = 4
        studenti_vizibili = []
        valid_set = set(valid_student_ids)
        display_student_ids = {str(sid): idx for idx, sid in enumerate(valid_student_ids, start=1)}

        for student in session_students or []:
            sid = str(student.get("student_id", "")).strip()
            if sid not in valid_set:
                continue

            severity = str(student.get("final_severity", "none") or "none").lower()
            alert_type = str(student.get("final_alert_type", "none") or "none").lower()
            if severity == "none" and alert_type == "none":
                continue

            studenti_vizibili.append(student)

        if not studenti_vizibili:
            ws.merge_cells("A4:H4")
            ws["A4"] = "No students required attention in this session."
            ws["A4"].alignment = Alignment(horizontal="center")
            ws["A4"].border = _border_subtire()
            _latime_automata(ws, min_width=14, max_width=40)
            return

        for student in studenti_vizibili:
            stare = self._stare_din_student(student)
            actiune = self._actiune_recomandata(student)
            observatie = self._humanize_reason(
                str(student.get("reason", "") or ""),
                float(student.get("attention_avg", 0.0) or 0.0),
                float(student.get("fatigue_avg", 0.0) or 0.0),
            )

            valori = [
                display_student_ids.get(str(student.get("student_id", "")).strip(), rand - 3),
                _rotunjeste(student.get("fatigue_avg", 0.0)),
                _rotunjeste(student.get("attention_avg", 0.0)),
                _rotunjeste(student.get("fatigue_max", 0.0)),
                _rotunjeste(student.get("attention_min", 0.0)),
                stare,
                actiune,
                observatie,
            ]

            for col_idx, value in enumerate(valori, start=1):
                cell = ws.cell(rand, col_idx, value)
                cell.border = _border_subtire()

                if col_idx == 1:
                    if isinstance(value, int):
                        cell.number_format = "0"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx in {2, 3, 4, 5}:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = "0.00"
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            fill = _fill_stare(stare)
            for col_idx in range(1, len(headers) + 1):
                ws.cell(rand, col_idx).fill = fill

            rand += 1

        widths = {
            "A": 14,
            "B": 14,
            "C": 16,
            "D": 14,
            "E": 14,
            "F": 14,
            "G": 32,
            "H": 42,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    def _construieste_sheet_date_trend(
        self,
        wb: Workbook,
        valid_student_ids: List[str],
        alert_student_ids: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        ws = wb.create_sheet("Ignore - Trend Data")
        ws.sheet_view.showGridLines = True

        student_ids = sorted(
            [str(x).strip() for x in valid_student_ids if str(x).strip()],
            key=lambda x: int(x) if x.isdigit() else x,
        )
        alert_ids = sorted(
            [str(x).strip() for x in (alert_student_ids or []) if str(x).strip() in set(student_ids)],
            key=lambda x: int(x) if x.isdigit() else x,
        )
        timestamps = [punct["timestamp_label"] for punct in self.trend_buffer]
        student_id_set = set(student_ids)
        alert_id_set = set(alert_ids)
        display_student_ids = {str(sid): idx for idx, sid in enumerate(student_ids, start=1)}

        ws.merge_cells("A1:N1")
        ws["A1"] = "Ignore - Raw data used to build the charts"
        _stil_titlu(ws["A1"])

        ws.merge_cells("A2:N3")
        ws["A2"] = (
            "This sheet stores the raw trend values used to generate the individual, group and alert-student charts. "
            "It can be ignored during normal review."
        )
        _stil_nota(ws["A2"])
        ws.row_dimensions[2].height = 26
        ws.row_dimensions[3].height = 26

        ws["A5"] = "Individual Fatigue Trend Data"
        _stil_titlu(ws["A5"])
        _stil_antet(ws.cell(6, 1, "Time"))
        for idx, student_id in enumerate(student_ids, start=2):
            _stil_antet(ws.cell(6, idx, f"Student {display_student_ids.get(str(student_id), idx - 1)}"))

        for row_idx, punct in enumerate(self.trend_buffer, start=7):
            ws.cell(row_idx, 1, punct["timestamp_label"]).border = _border_subtire()
            lookup = {
                str(item.get("student_id", "")).strip(): item
                for item in punct.get("studenti", [])
                if str(item.get("student_id", "")).strip() in student_id_set
            }
            for col_idx, student_id in enumerate(student_ids, start=2):
                cell = ws.cell(
                    row_idx,
                    col_idx,
                    None if lookup.get(student_id) is None else _rotunjeste(lookup[student_id].get("fatigue_pct", 0.0)),
                )
                cell.border = _border_subtire()
                if cell.value is not None:
                    cell.number_format = "0.00"

        att_title_row = len(self.trend_buffer) + 10
        ws.cell(att_title_row, 1, "Individual Attention Trend Data")
        _stil_titlu(ws.cell(att_title_row, 1))
        _stil_antet(ws.cell(att_title_row + 1, 1, "Time"))
        for idx, student_id in enumerate(student_ids, start=2):
            _stil_antet(ws.cell(att_title_row + 1, idx, f"Student {display_student_ids.get(str(student_id), idx - 1)}"))

        for row_idx, punct in enumerate(self.trend_buffer, start=att_title_row + 2):
            ws.cell(row_idx, 1, punct["timestamp_label"]).border = _border_subtire()
            lookup = {
                str(item.get("student_id", "")).strip(): item
                for item in punct.get("studenti", [])
                if str(item.get("student_id", "")).strip() in student_id_set
            }
            for col_idx, student_id in enumerate(student_ids, start=2):
                cell = ws.cell(
                    row_idx,
                    col_idx,
                    None if lookup.get(student_id) is None else _rotunjeste(lookup[student_id].get("attention_pct", 0.0)),
                )
                cell.border = _border_subtire()
                if cell.value is not None:
                    cell.number_format = "0.00"

        group_col = max(6, len(student_ids) + 4)
        group_header_row = 5
        ws.cell(group_header_row, group_col, "Group Trend Data")
        _stil_titlu(ws.cell(group_header_row, group_col))
        group_headers = ["Time", "Group fatigue", "Group attention", "Detected people", "Student alerts"]
        for idx, header in enumerate(group_headers, start=group_col):
            _stil_antet(ws.cell(group_header_row + 1, idx, header))

        for row_idx, punct in enumerate(self.trend_buffer, start=group_header_row + 2):
            values = [
                punct.get("timestamp_label", ""),
                _rotunjeste(punct.get("group_fatigue_pct", 0.0)),
                _rotunjeste(punct.get("group_attention_pct", 0.0)),
                int(punct.get("group_people_count", 0) or 0),
                int(punct.get("group_student_alert_count", 0) or 0),
            ]
            for idx, value in enumerate(values, start=group_col):
                cell = ws.cell(row_idx, idx, value)
                cell.border = _border_subtire()
                if idx > group_col:
                    cell.number_format = "0.00"

        alert_fatigue_header_row = att_title_row + len(self.trend_buffer) + 6
        ws.cell(alert_fatigue_header_row, 1, "Alert Student Fatigue Trend Data")
        _stil_titlu(ws.cell(alert_fatigue_header_row, 1))
        _stil_antet(ws.cell(alert_fatigue_header_row + 1, 1, "Time"))
        for idx, student_id in enumerate(alert_ids, start=2):
            _stil_antet(ws.cell(alert_fatigue_header_row + 1, idx, f"Student {display_student_ids.get(str(student_id), idx - 1)}"))

        for row_idx, punct in enumerate(self.trend_buffer, start=alert_fatigue_header_row + 2):
            ws.cell(row_idx, 1, punct["timestamp_label"]).border = _border_subtire()
            lookup = {
                str(item.get("student_id", "")).strip(): item
                for item in punct.get("studenti", [])
                if str(item.get("student_id", "")).strip() in alert_id_set
            }
            for col_idx, student_id in enumerate(alert_ids, start=2):
                cell = ws.cell(row_idx, col_idx, None if lookup.get(student_id) is None else _rotunjeste(lookup[student_id].get("fatigue_pct", 0.0)))
                cell.border = _border_subtire()
                if cell.value is not None:
                    cell.number_format = "0.00"

        alert_attention_header_row = alert_fatigue_header_row + len(self.trend_buffer) + 5
        ws.cell(alert_attention_header_row, 1, "Alert Student Attention Trend Data")
        _stil_titlu(ws.cell(alert_attention_header_row, 1))
        _stil_antet(ws.cell(alert_attention_header_row + 1, 1, "Time"))
        for idx, student_id in enumerate(alert_ids, start=2):
            _stil_antet(ws.cell(alert_attention_header_row + 1, idx, f"Student {display_student_ids.get(str(student_id), idx - 1)}"))

        for row_idx, punct in enumerate(self.trend_buffer, start=alert_attention_header_row + 2):
            ws.cell(row_idx, 1, punct["timestamp_label"]).border = _border_subtire()
            lookup = {
                str(item.get("student_id", "")).strip(): item
                for item in punct.get("studenti", [])
                if str(item.get("student_id", "")).strip() in alert_id_set
            }
            for col_idx, student_id in enumerate(alert_ids, start=2):
                cell = ws.cell(row_idx, col_idx, None if lookup.get(student_id) is None else _rotunjeste(lookup[student_id].get("attention_pct", 0.0)))
                cell.border = _border_subtire()
                if cell.value is not None:
                    cell.number_format = "0.00"

        _latime_automata(ws, min_width=14, max_width=22)
        return {
            "student_ids": student_ids,
            "alert_student_ids": alert_ids,
            "timestamps": timestamps,
            "attention_header_row": att_title_row + 1,
            "group_header_row": group_header_row + 1,
            "group_col": group_col,
            "alert_fatigue_header_row": alert_fatigue_header_row + 1,
            "alert_attention_header_row": alert_attention_header_row + 1,
        }

    def _construieste_sheet_grafice(
        self,
        wb: Workbook,
        trend_meta: Dict[str, Any],
        report_options: Optional[Dict[str, Any]] = None,
    ) -> None:
        report_options = report_options or {}
        report_type = str(report_options.get("report_type", "developer") or "developer").lower()
        developer_mode = report_type == "developer"
        chart_options = report_options.get("charts", {}) if isinstance(report_options.get("charts", {}), dict) else {}

        show_individual = developer_mode or bool(chart_options.get("individual", True))
        show_group = developer_mode or bool(chart_options.get("group", True))
        show_alerts = developer_mode or bool(chart_options.get("alerts", False))

        student_ids = list(trend_meta.get("student_ids", []) or [])
        alert_student_ids = list(trend_meta.get("alert_student_ids", []) or [])
        timestamps = list(trend_meta.get("timestamps", []) or [])
        attention_header_row = int(trend_meta.get("attention_header_row", 0) or 0)
        group_header_row = int(trend_meta.get("group_header_row", 0) or 0)
        group_col = int(trend_meta.get("group_col", 0) or 0)
        alert_fatigue_header_row = int(trend_meta.get("alert_fatigue_header_row", 0) or 0)
        alert_attention_header_row = int(trend_meta.get("alert_attention_header_row", 0) or 0)

        ws = wb.create_sheet("Student Trends")
        ws.sheet_view.showGridLines = True

        ws.merge_cells("A1:Z1")
        ws["A1"] = "Fatigue and Attention Trends"
        _stil_titlu(ws["A1"])

        ws.merge_cells("A2:Z3")
        ws["A2"] = (
            "Individual charts show how each included student evolved over time. Group charts show the class average. "
            "Alert-student charts show only the people who triggered attention or fatigue alerts."
        )
        _stil_nota(ws["A2"])
        ws.row_dimensions[2].height = 32
        ws.row_dimensions[3].height = 32

        if not timestamps:
            def add_no_data_session_block() -> None:
                ws.merge_cells("A5:L12")
                cell = ws["A5"]
                cell.value = (
                    "No data available for this session. "
                    "Please run the monitoring session for a longer period after calibration, then export the report again."
                )
                cell.font = Font(color="660000", bold=True, size=12)
                cell.fill = PatternFill("solid", fgColor="F4CCCC")
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                for row_idx in range(5, 13):
                    ws.row_dimensions[row_idx].height = 26
                    for col_idx in range(1, 13):
                        ws.cell(row_idx, col_idx).border = _border_subtire()
                        ws.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor="F4CCCC")
                for col_idx in range(1, 13):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 16

            add_no_data_session_block()
            return

        ws_date = wb["Ignore - Trend Data"]

        def make_chart(title: str, y_title: str, data_ref: Reference, cats_ref: Reference) -> LineChart:
            chart = LineChart()
            chart.title = title
            chart.y_axis.title = y_title
            chart.x_axis.title = "Time"
            chart.height = 10
            chart.width = 24
            chart.style = 10
            chart.legend.position = "t"
            try:
                chart.y_axis.scaling.min = 0
                chart.y_axis.scaling.max = 100
            except Exception:
                pass
            try:
                skip = max(1, len(timestamps) // 10)
                chart.x_axis.tickLblSkip = skip
                chart.x_axis.tickMarkSkip = skip
            except Exception:
                pass
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            for series in chart.series:
                try:
                    series.marker.symbol = "none"
                    series.graphicalProperties.line.width = 28000
                    series.smooth = False
                except Exception:
                    pass
            return chart

        def add_missing_block(anchor: str, title: str, message: str, width_cols: int = 11, height_rows: int = 7) -> None:
            start_cell = ws[anchor]
            start_row = start_cell.row
            start_col = start_cell.column
            end_col = min(start_col + width_cols, 26)
            end_row = start_row + height_rows

            title_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{start_row}"
            message_range = f"{get_column_letter(start_col)}{start_row + 1}:{get_column_letter(end_col)}{end_row}"

            ws.merge_cells(title_range)
            ws.merge_cells(message_range)

            title_cell = ws.cell(start_row, start_col)
            title_cell.value = f"⚠ {title}"
            title_cell.font = Font(color="FFFFFF", bold=True, size=12)
            title_cell.fill = PatternFill("solid", fgColor="C00000")
            title_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            title_cell.border = _border_subtire()

            msg_cell = ws.cell(start_row + 1, start_col)
            msg_cell.value = message
            msg_cell.font = Font(color="660000", bold=True, size=11)
            msg_cell.fill = PatternFill("solid", fgColor="F4CCCC")
            msg_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            msg_cell.border = _border_subtire()

            for row_idx in range(start_row, end_row + 1):
                ws.row_dimensions[row_idx].height = 26 if row_idx == start_row else 24
                for col_idx in range(start_col, end_col + 1):
                    cell = ws.cell(row_idx, col_idx)
                    cell.border = _border_subtire()
                    if row_idx > start_row:
                        cell.fill = PatternFill("solid", fgColor="F4CCCC")

            for col_idx in range(start_col, end_col + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 16

        if show_individual and not student_ids:
            add_missing_block(
                "A5",
                "Individual fatigue chart unavailable",
                "The option was selected in the UI, but the session does not contain enough valid student-level samples to build this chart. Run the session longer, keep at least one student visible after calibration, then export the report again.",
            )
            add_missing_block(
                "A28",
                "Individual attention chart unavailable",
                "The option was selected in the UI, but there are not enough valid attention samples per student. This is not a UI selection problem; the report has insufficient recorded data for this table/chart.",
            )

        if show_alerts and not alert_student_ids:
            add_missing_block(
                "A51",
                "Alert-student fatigue chart unavailable",
                "The option was selected in the UI, but no student triggered a fatigue/attention alert with enough valid samples. The report cannot create alert-student charts without recorded alert-student data.",
            )
            add_missing_block(
                "N51",
                "Alert-student attention chart unavailable",
                "The option was selected in the UI, but there are not enough alert-student attention samples. Export again after a session where at least one student remains in alert state long enough to be recorded.",
            )

        if show_individual and student_ids:
            data_fatigue = Reference(ws_date, min_col=2, max_col=1 + len(student_ids), min_row=6, max_row=6 + len(timestamps))
            cats_fatigue = Reference(ws_date, min_col=1, min_row=7, max_row=6 + len(timestamps))
            ws.add_chart(make_chart("Individual Fatigue Over Time", "Fatigue (%)", data_fatigue, cats_fatigue), "A5")

            attention_first_data_row = attention_header_row + 1
            attention_last_data_row = attention_header_row + len(timestamps)
            data_attention = Reference(ws_date, min_col=2, max_col=1 + len(student_ids), min_row=attention_header_row, max_row=attention_last_data_row)
            cats_attention = Reference(ws_date, min_col=1, min_row=attention_first_data_row, max_row=attention_last_data_row)
            ws.add_chart(make_chart("Individual Attention Over Time", "Attention (%)", data_attention, cats_attention), "A28")

        if show_group and group_col > 0:
            group_first_data_row = group_header_row + 1
            group_last_data_row = group_header_row + len(timestamps)
            cats_group = Reference(ws_date, min_col=group_col, min_row=group_first_data_row, max_row=group_last_data_row)
            data_group_fatigue = Reference(ws_date, min_col=group_col + 1, max_col=group_col + 1, min_row=group_header_row, max_row=group_last_data_row)
            data_group_attention = Reference(ws_date, min_col=group_col + 2, max_col=group_col + 2, min_row=group_header_row, max_row=group_last_data_row)
            ws.add_chart(make_chart("Group Fatigue Over Time", "Fatigue (%)", data_group_fatigue, cats_group), "N5")
            ws.add_chart(make_chart("Group Attention Over Time", "Attention (%)", data_group_attention, cats_group), "N28")

        if show_alerts and alert_student_ids:
            fatigue_first_data_row = alert_fatigue_header_row + 1
            fatigue_last_data_row = alert_fatigue_header_row + len(timestamps)
            attention_first_data_row = alert_attention_header_row + 1
            attention_last_data_row = alert_attention_header_row + len(timestamps)

            data_alert_fatigue = Reference(ws_date, min_col=2, max_col=1 + len(alert_student_ids), min_row=alert_fatigue_header_row, max_row=fatigue_last_data_row)
            cats_alert_fatigue = Reference(ws_date, min_col=1, min_row=fatigue_first_data_row, max_row=fatigue_last_data_row)
            data_alert_attention = Reference(ws_date, min_col=2, max_col=1 + len(alert_student_ids), min_row=alert_attention_header_row, max_row=attention_last_data_row)
            cats_alert_attention = Reference(ws_date, min_col=1, min_row=attention_first_data_row, max_row=attention_last_data_row)
            ws.add_chart(make_chart("Alert Students Fatigue Over Time", "Fatigue (%)", data_alert_fatigue, cats_alert_fatigue), "A51")
            ws.add_chart(make_chart("Alert Students Attention Over Time", "Attention (%)", data_alert_attention, cats_alert_attention), "N51")

    def exporta_raport_sesiune_xlsx(
        self,
        session_id: Optional[int],
        session_summary: Dict[str, Any],
        session_students: List[Dict[str, Any]],
        report_options: Optional[Dict[str, Any]] = None,
    ) -> str:
        report_options = _normalize_report_options(report_options)
        report_type = str(report_options.get("report_type", "teacher") or "teacher").lower()

        wb = Workbook()

        valid_student_ids = self._student_ids_valizi_pentru_raport(session_students)
        studenti_filtrati = self._filtreaza_studenti_sumar(session_students, valid_student_ids)
        alert_student_ids = [
            str(student.get("student_id", "")).strip()
            for student in studenti_filtrati
            if str(student.get("student_id", "")).strip()
            and (
                str(student.get("final_severity", "none") or "none").lower() != "none"
                or str(student.get("final_alert_type", "none") or "none").lower() != "none"
            )
        ]

        self._construieste_sheet_overview(wb, session_summary, report_options=report_options)
        self._construieste_sheet_studenti(wb, studenti_filtrati, valid_student_ids)
        trend_meta = self._construieste_sheet_date_trend(
            wb,
            valid_student_ids,
            alert_student_ids=alert_student_ids,
        )
        self._construieste_sheet_grafice(wb, trend_meta, report_options=report_options)

        desired_order = [
            "Session Overview",
            "Students Requiring Attention",
            "Student Trends",
            "Ignore - Trend Data",
        ]
        wb._sheets = [wb[name] for name in desired_order if name in wb.sheetnames]

        out_dir = Path(getattr(config, "SESSION_REPORTS_DIR", "reports"))
        out_dir.mkdir(parents=True, exist_ok=True)

        session_suffix = f"_{session_id}" if session_id is not None else ""
        ts_now = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = out_dir / f"raport_sesiune{session_suffix}_{report_type}_{ts_now}.xlsx"

        wb.save(out_path)

        if session_id is not None:
            update_session_raport_path(self.db_path, session_id, str(out_path))

        return str(out_path)
