import warnings
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from ultralytics import YOLO

try:
    import winsound
except Exception:
    winsound = None

import config
from attention import attentive_from_gaze, gaze_ratio
from fatigue import fatigue_percent, perclos_from_events

warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", message="SymbolDatabase.GetPrototype")
warnings.filterwarnings("ignore", module="google.protobuf.symbol_database")

LEFT_EYE = [33, 160, 158, 133, 153, 144]
RIGHT_EYE = [362, 385, 387, 263, 373, 380]

dataset_recorder_ref = None


def set_dataset_recorder(recorder):
    global dataset_recorder_ref
    dataset_recorder_ref = recorder


@dataclass
class PersonState:
    perclos_events: List[Tuple[float, bool]] = field(default_factory=list)
    eye_closed: bool = False
    eye_closed_start: Optional[float] = None

    last_fatigue_pct: float = 0.0
    last_attention_pct: float = 100.0

    fatigue_bad_since: Optional[float] = None
    attention_bad_since: Optional[float] = None

    fatigue_alert_active: bool = False
    attention_alert_active: bool = False

    fatigue_critical: bool = False
    attention_critical: bool = False

    last_seen_ts: float = 0.0
    valid_observations: int = 0


@dataclass
class EngineStats:
    faces: int = 0
    heads: int = 0
    class_avg_fatigue_pct: int = 0
    class_avg_attention_pct: int = 0

    alert_active: bool = False
    fatigue_alert_active: bool = False
    attention_alert_active: bool = False

    student_alerts: List[dict] = field(default_factory=list)
    new_alert_event: bool = False
    new_alert_kind: str = ""

    sound_alert_tick: bool = False
    decision_explanation: str = ""

    fps: float = 0.0
    alert_event_count: int = 0
    valid_observations: int = 0


@dataclass
class StaleTrack:
    raw_track_id: int
    persistent_id: int
    identity_id: Optional[int]
    bbox: Tuple[int, int, int, int]
    center: Tuple[float, float]
    saved_at: float
    person_state: PersonState


class HybridYoloDetector:
    def __init__(self):
        model_path = Path(getattr(config, "YOLO_MODEL_PATH", "best.pt"))
        conf = float(getattr(config, "YOLO_CONF", 0.45))
        imgsz = int(getattr(config, "YOLO_IMGSZ", 960))
        device = getattr(config, "YOLO_DEVICE", "cpu")

        if not model_path.exists():
            raise FileNotFoundError(f"YOLO model not found: {model_path}")

        self.model = YOLO(str(model_path))
        self.conf = conf
        self.imgsz = imgsz
        self.device = device

    @staticmethod
    def _clip_box(x1, y1, x2, y2, w, h):
        x1 = max(0, min(w - 1, int(x1)))
        y1 = max(0, min(h - 1, int(y1)))
        x2 = max(0, min(w - 1, int(x2)))
        y2 = max(0, min(h - 1, int(y2)))
        return x1, y1, x2, y2

    def _predict(self, frame_bgr):
        if frame_bgr is None or frame_bgr.size == 0:
            return []

        h, w = frame_bgr.shape[:2]

        result = self.model.predict(
            source=frame_bgr,
            conf=self.conf,
            imgsz=self.imgsz,
            device=self.device,
            verbose=False,
        )[0]

        items = []
        if result.boxes is None:
            return items

        for box in result.boxes:
            cls_id = int(box.cls[0].item())
            conf_score = float(box.conf[0].item()) if hasattr(box, "conf") else 0.0

            x1, y1, x2, y2 = map(int, box.xyxy[0].tolist())
            x1, y1, x2, y2 = self._clip_box(x1, y1, x2, y2, w, h)

            if x2 <= x1 + 1 or y2 <= y1 + 1:
                continue

            bw = x2 - x1
            bh = y2 - y1
            if bw < 8 or bh < 8:
                continue

            kind = None
            min_kind_conf = self.conf

            if cls_id == 0:
                kind = "face"
                min_kind_conf = float(getattr(config, "YOLO_FACE_CONF", self.conf))
            elif cls_id == 1:
                kind = "head"
                min_kind_conf = float(getattr(config, "YOLO_HEAD_CONF", self.conf))
            elif cls_id == 2:
                kind = "side_face"
                min_kind_conf = float(getattr(config, "YOLO_SIDEFACE_CONF", max(self.conf, 0.50)))

            if kind is None:
                continue

            if conf_score < min_kind_conf:
                continue

            items.append(
                {
                    "kind": kind,
                    "bbox": (x1, y1, x2, y2),
                    "conf": conf_score,
                }
            )

        return items

    def detect_heads(self, frame_bgr):
        return [x for x in self._predict(frame_bgr) if x["kind"] == "head"]

    def detect_faces(self, frame_bgr):
        return [x for x in self._predict(frame_bgr) if x["kind"] in ("face", "side_face")]


def lm_xy(face_lms, idx, w, h):
    p = face_lms.landmark[idx]
    return np.array([p.x * w, p.y * h], dtype=np.float32)


def expand_bbox(box, w_img, h_img, scale=1.03):
    if box is None:
        return None

    x1, y1, x2, y2 = box
    cx = (x1 + x2) * 0.5
    cy = (y1 + y2) * 0.5
    bw = (x2 - x1) * scale
    bh = (y2 - y1) * scale

    nx1 = int(max(0, cx - bw * 0.5))
    ny1 = int(max(0, cy - bh * 0.5))
    nx2 = int(min(w_img - 1, cx + bw * 0.5))
    ny2 = int(min(h_img - 1, cy + bh * 0.5))

    if nx2 <= nx1 + 1 or ny2 <= ny1 + 1:
        return None

    return nx1, ny1, nx2, ny2


def shrink_box(box, w_img, h_img, sx=0.92, sy=0.92):
    if box is None:
        return None

    x1, y1, x2, y2 = box
    cx = (x1 + x2) * 0.5
    cy = (y1 + y2) * 0.5
    bw = (x2 - x1) * sx
    bh = (y2 - y1) * sy

    nx1 = int(max(0, cx - bw * 0.5))
    ny1 = int(max(0, cy - bh * 0.5))
    nx2 = int(min(w_img - 1, cx + bw * 0.5))
    ny2 = int(min(h_img - 1, cy + bh * 0.5))

    if nx2 <= nx1 + 1 or ny2 <= ny1 + 1:
        return None

    return nx1, ny1, nx2, ny2


def infer_head_box_from_face(face_box, w_img, h_img):
    if face_box is None:
        return None

    x1, y1, x2, y2 = face_box
    fw = x2 - x1
    fh = y2 - y1

    if fw <= 1 or fh <= 1:
        return None

    cx = (x1 + x2) * 0.5
    face_bottom = y2

    head_w = fw * float(getattr(config, "INFER_HEAD_FROM_FACE_W_SCALE", 1.22))
    head_h = fh * float(getattr(config, "INFER_HEAD_FROM_FACE_H_SCALE", 1.38))

    top_extra = fh * float(getattr(config, "INFER_HEAD_FROM_FACE_TOP_EXTRA", 0.28))
    bottom_extra = fh * float(getattr(config, "INFER_HEAD_FROM_FACE_BOTTOM_EXTRA", 0.10))

    nx1 = int(max(0, cx - head_w * 0.5))
    nx2 = int(min(w_img - 1, cx + head_w * 0.5))
    ny1 = int(max(0, y1 - top_extra))
    ny2 = int(min(h_img - 1, y2 + bottom_extra))

    target_h = int(max(1, head_h))
    cur_h = ny2 - ny1
    if cur_h > target_h:
        extra = cur_h - target_h
        ny1 += extra // 2
        ny2 -= extra - (extra // 2)

    if nx2 <= nx1 + 1 or ny2 <= ny1 + 1:
        return None

    box = (nx1, ny1, nx2, ny2)
    if not looks_like_head_box(box, w_img, h_img):
        return None

    return box


def eye_aspect_ratio(face_lms, eye_idx, w, h) -> float:
    p1 = lm_xy(face_lms, eye_idx[0], w, h)
    p2 = lm_xy(face_lms, eye_idx[1], w, h)
    p3 = lm_xy(face_lms, eye_idx[2], w, h)
    p4 = lm_xy(face_lms, eye_idx[3], w, h)
    p5 = lm_xy(face_lms, eye_idx[4], w, h)
    p6 = lm_xy(face_lms, eye_idx[5], w, h)

    v1 = np.linalg.norm(p2 - p6)
    v2 = np.linalg.norm(p3 - p5)
    h1 = np.linalg.norm(p1 - p4)

    return float((v1 + v2) / (2.0 * h1 + 1e-9))


def bbox_area(box):
    x1, y1, x2, y2 = box
    return max(0, x2 - x1) * max(0, y2 - y1)


def bbox_center(box):
    x1, y1, x2, y2 = box
    return ((x1 + x2) * 0.5, (y1 + y2) * 0.5)


def box_contains_point(box, px, py):
    x1, y1, x2, y2 = box
    return x1 <= px <= x2 and y1 <= py <= y2


def iou(box_a, box_b):
    ax1, ay1, ax2, ay2 = box_a
    bx1, by1, bx2, by2 = box_b

    ix1 = max(ax1, bx1)
    iy1 = max(ay1, by1)
    ix2 = min(ax2, bx2)
    iy2 = min(ay2, by2)

    iw = max(0, ix2 - ix1)
    ih = max(0, iy2 - iy1)
    inter = iw * ih

    if inter <= 0:
        return 0.0

    area_a = bbox_area(box_a)
    area_b = bbox_area(box_b)
    return inter / (area_a + area_b - inter + 1e-9)


def normalize_xyxy(box, w, h):
    x1, y1, x2, y2 = box
    return [x1 / w, y1 / h, x2 / w, y2 / h]


def nms_xyxy(boxes, iou_thresh=0.45):
    if not boxes:
        return []

    ordered = sorted(boxes, key=bbox_area, reverse=True)
    kept = []

    for box in ordered:
        ok = True
        for prev in kept:
            if iou(box, prev) >= iou_thresh:
                ok = False
                break
        if ok:
            kept.append(box)

    return kept


def nms_dict_boxes(items, iou_thresh=0.45):
    if not items:
        return []

    ordered = sorted(
        items,
        key=lambda x: (x.get("conf", 0.0), bbox_area(x["bbox"])),
        reverse=True,
    )
    kept = []

    for item in ordered:
        ok = True
        for prev in kept:
            if iou(item["bbox"], prev["bbox"]) >= iou_thresh:
                ok = False
                break
        if ok:
            kept.append(item)

    return kept


def looks_like_face_box(box, w_img, h_img):
    x1, y1, x2, y2 = box
    bw = x2 - x1
    bh = y2 - y1

    if bw < getattr(config, "FACE_BOX_MIN_SIZE", 14) or bh < getattr(config, "FACE_BOX_MIN_SIZE", 14):
        return False

    if bw > getattr(config, "FACE_BOX_MAX_W_RATIO", 0.90) * w_img:
        return False

    if bh > getattr(config, "FACE_BOX_MAX_H_RATIO", 0.90) * h_img:
        return False

    ratio = bw / (bh + 1e-9)
    if ratio < getattr(config, "FACE_BOX_MIN_ASPECT", 0.40):
        return False
    if ratio > getattr(config, "FACE_BOX_MAX_ASPECT", 2.20):
        return False

    return True


def looks_like_head_box(box, w_img, h_img):
    x1, y1, x2, y2 = box
    bw = x2 - x1
    bh = y2 - y1

    if bw < getattr(config, "HEAD_BOX_MIN_SIZE", 16) or bh < getattr(config, "HEAD_BOX_MIN_SIZE", 16):
        return False

    if bw > getattr(config, "HEAD_BOX_MAX_W_RATIO", 0.95) * w_img:
        return False

    if bh > getattr(config, "HEAD_BOX_MAX_H_RATIO", 0.99) * h_img:
        return False

    ratio = bw / (bh + 1e-9)
    if ratio < getattr(config, "HEAD_BOX_MIN_ASPECT", 0.22):
        return False
    if ratio > getattr(config, "HEAD_BOX_MAX_ASPECT", 2.40):
        return False

    return True


def face_inside_head_score(face_box, head_box):
    fc_x, fc_y = bbox_center(face_box)
    hc_x, hc_y = bbox_center(head_box)

    if not box_contains_point(head_box, fc_x, fc_y):
        return -1e9

    overlap = iou(face_box, head_box)
    face_area = bbox_area(face_box)
    head_area = bbox_area(head_box)
    area_ratio = face_area / (head_area + 1e-9)

    if area_ratio < 0.05 or area_ratio > 0.82:
        return -1e9

    dx = abs(fc_x - hc_x) / max(1.0, head_box[2] - head_box[0])
    dy = abs(fc_y - hc_y) / max(1.0, head_box[3] - head_box[1])

    score = 0.0
    score += overlap * 6000.0
    score += face_area
    score += max(0.0, 1.0 - dx) * 800.0
    score += max(0.0, 1.0 - dy) * 600.0

    top_bonus = max(0.0, 1.0 - ((fc_y - head_box[1]) / max(1.0, head_box[3] - head_box[1])))
    score += top_bonus * 400.0
    return score


def box_key(box, q=20):
    x1, y1, x2, y2 = box
    return (
        int(x1 // q),
        int(y1 // q),
        int(x2 // q),
        int(y2 // q),
    )


def valid_face_area(box):
    area = bbox_area(box)
    return area >= getattr(config, "FACE_BOX_MIN_AREA", 1200)


def valid_head_area(box):
    area = bbox_area(box)
    return area >= getattr(config, "HEAD_BOX_MIN_AREA", 700)


def clone_person_state(st: PersonState) -> PersonState:
    return PersonState(
        perclos_events=list(st.perclos_events),
        eye_closed=bool(st.eye_closed),
        eye_closed_start=st.eye_closed_start,
        last_fatigue_pct=float(st.last_fatigue_pct),
        last_attention_pct=float(st.last_attention_pct),
        fatigue_bad_since=st.fatigue_bad_since,
        attention_bad_since=st.attention_bad_since,
        fatigue_alert_active=bool(st.fatigue_alert_active),
        attention_alert_active=bool(st.attention_alert_active),
        fatigue_critical=bool(st.fatigue_critical),
        attention_critical=bool(st.attention_critical),
        last_seen_ts=float(st.last_seen_ts),
        valid_observations=int(st.valid_observations),
    )


def update_boolean_alert_with_hysteresis(
    value: float,
    is_high_alert: bool,
    on_threshold: float,
    off_threshold: float,
    hold_s: float,
    active: bool,
    bad_since: Optional[float],
    now: float,
):
    if is_high_alert:
        bad_now = value >= on_threshold
        good_now = value < off_threshold
    else:
        bad_now = value <= on_threshold
        good_now = value > off_threshold

    if active:
        if good_now:
            active = False
            bad_since = None
    else:
        if bad_now:
            if bad_since is None:
                bad_since = now
            elif (now - bad_since) >= hold_s:
                active = True
        else:
            bad_since = None

    return active, bad_since


def student_alert_severity(fatigue_pct: float, attention_pct: float) -> str:
    if (
        fatigue_pct >= float(getattr(config, "ALERT_STUDENT_FATIGUE_CRITICAL", 70))
        or attention_pct <= float(getattr(config, "ALERT_STUDENT_ATTENTION_CRITICAL", 30))
    ):
        return "critical"

    if (
        fatigue_pct >= float(getattr(config, "ALERT_STUDENT_FATIGUE_ON", 60))
        or attention_pct <= float(getattr(config, "ALERT_STUDENT_ATTENTION_ON", 40))
    ):
        return "warning"

    return "ok"


def build_student_alert_message(student_id: int, st: PersonState) -> str:
    if st.attention_alert_active and st.fatigue_alert_active:
        if st.attention_critical or st.fatigue_critical:
            return f"Student {student_id} is very inattentive and very fatigued"
        return f"Student {student_id} is inattentive and fatigued"

    if st.attention_alert_active:
        if st.attention_critical:
            return f"Student {student_id} is very inattentive"
        return f"Student {student_id} is inattentive"

    if st.fatigue_alert_active:
        if st.fatigue_critical:
            return f"Student {student_id} is very fatigued"
        return f"Student {student_id} is fatigued"

    return ""


def compute_student_alerts(people: Dict[int, PersonState], now: float):
    alerts = []
    new_alert_event = False
    new_alert_kind = ""

    for student_id, st in people.items():
        if (now - st.last_seen_ts) > max(1.0, float(getattr(config, "TRACK_TTL_S", 3.0))):
            continue

        if int(st.valid_observations) < int(getattr(config, "METRIC_READY_MIN_FRAMES", 20)):
            continue

        prev_fatigue = bool(st.fatigue_alert_active)
        prev_attention = bool(st.attention_alert_active)

        st.fatigue_alert_active, st.fatigue_bad_since = update_boolean_alert_with_hysteresis(
            value=float(st.last_fatigue_pct),
            is_high_alert=True,
            on_threshold=float(getattr(config, "ALERT_STUDENT_FATIGUE_ON", 60)),
            off_threshold=float(getattr(config, "ALERT_STUDENT_FATIGUE_OFF", 52)),
            hold_s=float(getattr(config, "ALERT_STUDENT_HOLD_S", 4.0)),
            active=bool(st.fatigue_alert_active),
            bad_since=st.fatigue_bad_since,
            now=now,
        )

        st.attention_alert_active, st.attention_bad_since = update_boolean_alert_with_hysteresis(
            value=float(st.last_attention_pct),
            is_high_alert=False,
            on_threshold=float(getattr(config, "ALERT_STUDENT_ATTENTION_ON", 40)),
            off_threshold=float(getattr(config, "ALERT_STUDENT_ATTENTION_OFF", 50)),
            hold_s=float(getattr(config, "ALERT_STUDENT_HOLD_S", 4.0)),
            active=bool(st.attention_alert_active),
            bad_since=st.attention_bad_since,
            now=now,
        )

        st.fatigue_critical = float(st.last_fatigue_pct) >= float(
            getattr(config, "ALERT_STUDENT_FATIGUE_CRITICAL", 70)
        )
        st.attention_critical = float(st.last_attention_pct) <= float(
            getattr(config, "ALERT_STUDENT_ATTENTION_CRITICAL", 30)
        )

        fatigue_new = (not prev_fatigue) and st.fatigue_alert_active
        attention_new = (not prev_attention) and st.attention_alert_active

        if fatigue_new or attention_new:
            new_alert_event = True
            if st.attention_critical or st.fatigue_critical:
                new_alert_kind = "student_critical"
            elif st.attention_alert_active and st.fatigue_alert_active:
                new_alert_kind = "student_both"
            elif st.attention_alert_active:
                new_alert_kind = "student_attention"
            elif st.fatigue_alert_active:
                new_alert_kind = "student_fatigue"

        if st.fatigue_alert_active or st.attention_alert_active:
            severity = student_alert_severity(float(st.last_fatigue_pct), float(st.last_attention_pct))

            recommended_action = ""
            if severity == "critical":
                recommended_action = "Check this student immediately"
            elif st.attention_alert_active and st.fatigue_alert_active:
                recommended_action = "Consider checking engagement and fatigue"
            elif st.attention_alert_active:
                recommended_action = "Consider re-engaging this student"
            elif st.fatigue_alert_active:
                recommended_action = "Consider a short break or check-in"

            alerts.append(
                {
                    "student_id": int(student_id),
                    "scope": "student",
                    "fatigue_pct": round(float(st.last_fatigue_pct), 1),
                    "attention_pct": round(float(st.last_attention_pct), 1),
                    "fatigue_alert_active": bool(st.fatigue_alert_active),
                    "attention_alert_active": bool(st.attention_alert_active),
                    "fatigue_critical": bool(st.fatigue_critical),
                    "attention_critical": bool(st.attention_critical),
                    "severity": severity,
                    "message": build_student_alert_message(int(student_id), st),
                    "recommended_action": recommended_action,
                    "valid_observations": int(st.valid_observations),
                }
            )

    alerts.sort(
        key=lambda x: (
            0 if x["severity"] == "critical" else 1,
            -(float(x["fatigue_pct"]) + (100.0 - float(x["attention_pct"]))),
        )
    )

    max_items = int(getattr(config, "MAX_STUDENT_ALERTS_IN_UI", 5))
    if max_items < 1:
        max_items = 1

    return alerts[:max_items], new_alert_event, new_alert_kind


def compute_class_alerts(
    class_avg_fatigue: int,
    class_avg_attention: int,
    active_student_count: int,
    now: float,
    class_fatigue_alert_active: bool,
    class_attention_alert_active: bool,
    class_fatigue_bad_since: Optional[float],
    class_attention_bad_since: Optional[float],
):
    min_students = int(getattr(config, "MIN_ACTIVE_STUDENTS_FOR_CLASS_ALERT", 2))
    if min_students < 1:
        min_students = 1

    enough_students = active_student_count >= min_students

    prev_fatigue = bool(class_fatigue_alert_active)
    prev_attention = bool(class_attention_alert_active)

    if enough_students:
        class_fatigue_alert_active, class_fatigue_bad_since = update_boolean_alert_with_hysteresis(
            value=float(class_avg_fatigue),
            is_high_alert=True,
            on_threshold=float(getattr(config, "ALERT_CLASS_FATIGUE_ON", getattr(config, "ALERT_FATIGUE_PCT", 50))),
            off_threshold=float(getattr(config, "ALERT_CLASS_FATIGUE_OFF", 48)),
            hold_s=float(getattr(config, "ALERT_CLASS_HOLD_S", 6.0)),
            active=bool(class_fatigue_alert_active),
            bad_since=class_fatigue_bad_since,
            now=now,
        )

        class_attention_alert_active, class_attention_bad_since = update_boolean_alert_with_hysteresis(
            value=float(class_avg_attention),
            is_high_alert=False,
            on_threshold=float(getattr(config, "ALERT_CLASS_ATTENTION_ON", getattr(config, "ALERT_ATTENTION_MIN_PCT", 50))),
            off_threshold=float(getattr(config, "ALERT_CLASS_ATTENTION_OFF", 52)),
            hold_s=float(getattr(config, "ALERT_CLASS_HOLD_S", 6.0)),
            active=bool(class_attention_alert_active),
            bad_since=class_attention_bad_since,
            now=now,
        )
    else:
        class_fatigue_alert_active = False
        class_attention_alert_active = False
        class_fatigue_bad_since = None
        class_attention_bad_since = None

    fatigue_new = (not prev_fatigue) and class_fatigue_alert_active
    attention_new = (not prev_attention) and class_attention_alert_active
    class_alert_active = bool(class_fatigue_alert_active or class_attention_alert_active)

    reasons = []
    if class_alert_active:
        if class_attention_alert_active:
            reasons.append(
                f"class attention low ({class_avg_attention}% <= ON {int(getattr(config, 'ALERT_CLASS_ATTENTION_ON', getattr(config, 'ALERT_ATTENTION_MIN_PCT', 50)))})"
            )
        if class_fatigue_alert_active:
            reasons.append(
                f"class fatigue high ({class_avg_fatigue}% >= ON {int(getattr(config, 'ALERT_CLASS_FATIGUE_ON', getattr(config, 'ALERT_FATIGUE_PCT', 50)))})"
            )
    else:
        reasons.append(
            f"no class alert: fatigue={class_avg_fatigue}%, attention={class_avg_attention}%, active_students={active_student_count}"
        )
        if not enough_students:
            reasons.append(f"minimum students not reached ({active_student_count}/{min_students})")

    new_kind = ""
    if fatigue_new and attention_new:
        new_kind = "class_both"
    elif fatigue_new:
        new_kind = "class_fatigue"
    elif attention_new:
        new_kind = "class_attention"

    return (
        class_alert_active,
        bool(class_fatigue_alert_active),
        bool(class_attention_alert_active),
        bool(fatigue_new or attention_new),
        new_kind,
        "; ".join(reasons),
        class_fatigue_bad_since,
        class_attention_bad_since,
    )


def maybe_trigger_sound_alert(now: float, should_trigger: bool, last_any_alert_sound_ts: float):
    if not should_trigger:
        return False, last_any_alert_sound_ts

    if not bool(getattr(config, "ENABLE_SOUND_ALERT", True)):
        return False, last_any_alert_sound_ts

    min_interval = float(getattr(config, "SOUND_ALERT_MIN_INTERVAL_S", 8.0))
    if (now - float(last_any_alert_sound_ts)) < min_interval:
        return False, last_any_alert_sound_ts

    played = False

    try:
        sound_path = getattr(config, "ALERT_SOUND_FILE", None)
        if sound_path:
            sound_path = Path(sound_path)
            if sound_path.exists() and winsound is not None:
                winsound.PlaySound(str(sound_path), winsound.SND_FILENAME | winsound.SND_ASYNC)
                played = True
        elif winsound is not None:
            freq = int(getattr(config, "ALERT_BEEP_FREQ", 880))
            duration_ms = int(getattr(config, "ALERT_BEEP_DURATION_MS", 180))
            winsound.Beep(freq, duration_ms)
            played = True
    except Exception:
        played = False

    if played:
        last_any_alert_sound_ts = now

    return played, last_any_alert_sound_ts


def export_session_raport_xlsx(
    summary: dict,
    session_students: Optional[List[dict]] = None,
) -> str:
    if not summary.get("has_data"):
        raise ValueError("No session data available")

    if session_students is None:
        session_students = []

    reports_dir = Path(getattr(config, "SESSION_RAPORTS_DIR", getattr(config, "SESSION_REPORTS_DIR", "rapoarte")))
    reports_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = reports_dir / f"session_raport_{ts}.xlsx"

    wb = Workbook()

    # styles
    title_fill = PatternFill("solid", fgColor="6F558C")
    section_fill = PatternFill("solid", fgColor="DCE6F1")
    header_fill = PatternFill("solid", fgColor="26B6DE")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    critical_fill = PatternFill("solid", fgColor="F4CCCC")
    ok_fill = PatternFill("solid", fgColor="E2F0D9")

    white_bold = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)

    thin = Side(style="thin", color="7F7F7F")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # =========================================================
    # SHEET 1 - SESSION OVERVIEW
    # =========================================================
    ws = wb.active
    ws.title = "Session Overview"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22

    ws["A1"] = "Class Monitor - Session Overview"
    ws["A1"].fill = title_fill
    ws["A1"].font = white_bold
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = border
    for c in range(2, 5):
        ws.cell(row=1, column=c).border = border
    ws.merge_cells("A1:D1")

    started_str = datetime.fromtimestamp(summary["started_at"]).strftime("%Y-%m-%d %H:%M:%S")
    stopped_str = datetime.fromtimestamp(summary["stopped_at"]).strftime("%Y-%m-%d %H:%M:%S")
    duration_min = round(float(summary.get("duration_s", 0.0)) / 60.0, 2)

    overview_rows = [
        ("Started at", started_str),
        ("Stopped at", stopped_str),
        ("Duration (minutes)", duration_min),
        ("Samples analyzed", int(summary.get("samples", 0))),
        ("Unique faces seen", int(summary.get("unique_faces_seen", 0))),
        ("Max faces seen", int(summary.get("max_faces_seen", 0))),
        ("Max heads seen", int(summary.get("max_heads_seen", 0))),
        ("Avg faces", round(float(summary.get("avg_faces", 0.0)), 2)),
        ("Avg heads", round(float(summary.get("avg_heads", 0.0)), 2)),
        ("Avg fatigue (%)", round(float(summary.get("avg_fatigue", 0.0)), 2)),
        ("Avg attention (%)", round(float(summary.get("avg_attention", 0.0)), 2)),
        ("Alert events", int(summary.get("alert_event_count", 0))),
    ]

    row = 3
    for label, value in overview_rows:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"A{row}"].fill = section_fill
        ws[f"A{row}"].font = bold
        ws[f"A{row}"].border = border
        ws[f"B{row}"].border = border
        row += 1

    flagged_students = [
        s for s in session_students
        if str(s.get("final_severity", "none")).lower() != "none"
    ]
    critical_count = sum(
        1 for s in flagged_students
        if str(s.get("final_severity", "none")).lower() == "critical"
    )
    warning_count = sum(
        1 for s in flagged_students
        if str(s.get("final_severity", "none")).lower() == "warning"
    )

    ws["D3"] = "Session Interpretation"
    ws["D3"].fill = section_fill
    ws["D3"].font = bold
    ws["D3"].border = border

    interpretation_rows = [
        ("Problematic students", len(flagged_students)),
        ("Critical students", critical_count),
        ("Warning students", warning_count),
    ]

    row = 4
    for label, value in interpretation_rows:
        ws[f"C{row}"] = label
        ws[f"D{row}"] = value
        ws[f"C{row}"].fill = section_fill
        ws[f"C{row}"].font = bold
        ws[f"C{row}"].border = border
        ws[f"D{row}"].border = border
        row += 1

    ws["C8"] = "Summary"
    ws["C8"].fill = section_fill
    ws["C8"].font = bold
    ws["C8"].border = border
    ws["D8"].border = border

    summary_text = (
        f"Session duration: {duration_min} min | "
        f"Average fatigue: {round(float(summary.get('avg_fatigue', 0.0)), 2)}% | "
        f"Average attention: {round(float(summary.get('avg_attention', 0.0)), 2)}% | "
        f"Problematic students: {len(flagged_students)}"
    )

    ws["C9"] = summary_text
    ws["C9"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["C9"].border = border
    ws["D9"].border = border
    ws.merge_cells("C9:D12")

    # =========================================================
    # SHEET 2 - STUDENT ISSUES
    # =========================================================
    ws2 = wb.create_sheet("Student Issues")

    widths = {
        "A": 14,
        "B": 14,
        "C": 16,
        "D": 14,
        "E": 16,
        "F": 18,
        "G": 12,
        "H": 34,
        "I": 44,
    }
    for col, width in widths.items():
        ws2.column_dimensions[col].width = width

    ws2["A1"] = "Students with Issues"
    ws2["A1"].fill = title_fill
    ws2["A1"].font = white_bold
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["A1"].border = border
    for c in range(2, 10):
        ws2.cell(row=1, column=c).border = border
    ws2.merge_cells("A1:I1")

    headers = [
        "Student ID",
        "Fatigue Avg",
        "Attention Avg",
        "Fatigue Max",
        "Attention Min",
        "Alert Type",
        "Severity",
        "Decision",
        "Reason",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = white_bold
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if not flagged_students:
        ws2["A4"] = "No students with flagged issues detected in this session."
        ws2["A4"].border = border
        ws2["A4"].alignment = Alignment(horizontal="center", vertical="center")
        for c in range(2, 10):
            ws2.cell(row=4, column=c).border = border
        ws2.merge_cells("A4:I4")
    else:
        row_idx = 4
        for student in flagged_students:
            values = [
                student.get("student_id", ""),
                round(float(student.get("fatigue_avg", 0.0)), 2),
                round(float(student.get("attention_avg", 0.0)), 2),
                round(float(student.get("fatigue_max", 0.0)), 2),
                round(float(student.get("attention_min", 0.0)), 2),
                student.get("final_alert_type", "none"),
                student.get("final_severity", "none"),
                student.get("decision", ""),
                student.get("reason", ""),
            ]

            severity = str(student.get("final_severity", "none")).lower()
            row_fill = ok_fill
            if severity == "critical":
                row_fill = critical_fill
            elif severity == "warning":
                row_fill = warning_fill

            for col_idx, value in enumerate(values, start=1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.fill = row_fill
                cell.alignment = Alignment(
                    horizontal="center" if col_idx <= 7 else "left",
                    vertical="center",
                    wrap_text=True,
                )
            row_idx += 1

    wb.save(out_path)
    return str(out_path)


# Backward-compatible alias for older imports.
def export_session_report_xlsx(summary: dict, session_students: Optional[List[dict]] = None) -> str:
    return export_session_raport_xlsx(summary=summary, session_students=session_students)


def is_box_in_reasonable_vertical_zone(box, h_img):
    _cx, cy = bbox_center(box)
    min_y_ratio = float(getattr(config, "FACE_CENTER_MIN_Y_RATIO", 0.08))
    max_y_ratio = float(getattr(config, "FACE_CENTER_MAX_Y_RATIO", 0.92))
    return (cy >= min_y_ratio * h_img) and (cy <= max_y_ratio * h_img)


def is_face_candidate_stable(box, w_img, h_img):
    if not looks_like_face_box(box, w_img, h_img):
        return False

    area = bbox_area(box)
    if area < int(getattr(config, "FACE_BOX_MIN_AREA", 900)):
        return False

    max_area_ratio = float(getattr(config, "FACE_BOX_MAX_AREA_RATIO", 0.16))
    if area > (w_img * h_img * max_area_ratio):
        return False

    if not is_box_in_reasonable_vertical_zone(box, h_img):
        return False

    x1, _y1, x2, _y2 = box
    margin_ratio = float(getattr(config, "FACE_EDGE_MARGIN_RATIO", 0.01))
    if x1 <= margin_ratio * w_img or x2 >= (1.0 - margin_ratio) * w_img:
        return False

    return True


def is_head_candidate_stable(box, w_img, h_img):
    if not looks_like_head_box(box, w_img, h_img):
        return False

    area = bbox_area(box)
    if area < int(getattr(config, "HEAD_BOX_MIN_AREA", 700)):
        return False

    max_area_ratio = float(getattr(config, "HEAD_BOX_MAX_AREA_RATIO", 0.24))
    if area > (w_img * h_img * max_area_ratio):
        return False

    x1, _y1, x2, _y2 = box
    margin_ratio = float(getattr(config, "HEAD_EDGE_MARGIN_RATIO", 0.01))
    if x1 <= margin_ratio * w_img or x2 >= (1.0 - margin_ratio) * w_img:
        return False

    _cx, cy = bbox_center(box)
    if cy < float(getattr(config, "HEAD_CENTER_MIN_Y_RATIO", 0.05)) * h_img:
        return False
    if cy > float(getattr(config, "HEAD_CENTER_MAX_Y_RATIO", 0.95)) * h_img:
        return False

    return True


def mesh_confirms_face_box(mesh, frame_bgr: np.ndarray, face_box: Tuple[int, int, int, int], scale: float = 1.16):
    h_img, w_img = frame_bgr.shape[:2]

    if not is_face_candidate_stable(face_box, w_img, h_img):
        return False, None, None, None

    mesh_box = expand_bbox(face_box, w_img, h_img, scale=scale)
    if mesh_box is None:
        mesh_box = face_box

    mx1, my1, mx2, my2 = mesh_box
    mesh_face = frame_bgr[my1:my2, mx1:mx2]
    if mesh_face.size == 0:
        return False, None, None, None

    mesh_h, mesh_w = mesh_face.shape[:2]
    if mesh_w < int(getattr(config, "FACE_MESH_MIN_W", 24)) or mesh_h < int(getattr(config, "FACE_MESH_MIN_H", 24)):
        return False, None, None, None

    rgb = np.ascontiguousarray(mesh_face[:, :, ::-1])
    res = mesh.process(rgb)

    if not res or not res.multi_face_landmarks:
        return False, None, None, None

    lms = res.multi_face_landmarks[0]

    try:
        ear_l = eye_aspect_ratio(lms, LEFT_EYE, mesh_w, mesh_h)
        ear_r = eye_aspect_ratio(lms, RIGHT_EYE, mesh_w, mesh_h)
        ear = (ear_l + ear_r) / 2.0
    except Exception:
        return False, None, None, None

    if not np.isfinite(ear):
        return False, None, None, None

    return True, lms, mesh_w, mesh_h


def filter_mp_faces(face_detector, frame_bgr: np.ndarray):
    h_img, w_img = frame_bgr.shape[:2]
    boxes = face_detector.detect(frame_bgr, detect_width=config.DETECT_WIDTH)

    valid = []
    for b in boxes:
        if not looks_like_face_box(b, w_img, h_img):
            continue
        if not valid_face_area(b):
            continue

        _cx, cy = bbox_center(b)
        if cy < getattr(config, "FACE_CENTER_MIN_Y_RATIO", 0.08) * h_img:
            continue
        if cy > getattr(config, "FACE_CENTER_MAX_Y_RATIO", 0.92) * h_img:
            continue

        valid.append(b)

    return nms_xyxy(valid, iou_thresh=0.45)


def detect_yolo_faces(yolo_detector, frame_bgr: np.ndarray):
    return yolo_detector.detect_faces(frame_bgr)


def filter_yolo_faces(yolo_detector, frame_bgr: np.ndarray):
    items = detect_yolo_faces(yolo_detector, frame_bgr)
    if not items:
        return []

    h_img, w_img = frame_bgr.shape[:2]
    valid = []

    for x in items:
        box = x["bbox"]

        if not is_face_candidate_stable(box, w_img, h_img):
            continue

        cx, cy = bbox_center(box)

        if cy < 0.05 * h_img:
            continue
        if cy > 0.95 * h_img:
            continue
        if cx < 0.01 * w_img:
            continue
        if cx > 0.99 * w_img:
            continue

        valid.append(x)

    return nms_dict_boxes(valid, iou_thresh=0.45)


def filter_heads(yolo_detector, head_detector, frame_bgr: np.ndarray):
    h_img, w_img = frame_bgr.shape[:2]

    yolo_heads = yolo_detector.detect_heads(frame_bgr)
    yolo_heads = [
        x for x in yolo_heads
        if is_head_candidate_stable(x["bbox"], w_img, h_img)
        and valid_head_area(x["bbox"])
    ]

    if head_detector.enabled:
        api_heads = head_detector.detect(frame_bgr)
        api_heads = [
            {"kind": "head", "bbox": b, "conf": 0.5}
            for b in api_heads
            if is_head_candidate_stable(b, w_img, h_img)
            and valid_head_area(b)
        ]
        merged = yolo_heads + api_heads
    else:
        merged = yolo_heads

    return nms_dict_boxes(merged, iou_thresh=0.35)


def best_face_inside_head(
    face_detector,
    yolo_faces: List[dict],
    frame_bgr: np.ndarray,
    head_box: Tuple[int, int, int, int],
    mp_faces: List[Tuple[int, int, int, int]],
):
    h_img, w_img = frame_bgr.shape[:2]

    best_box = None
    best_face_kind = "face"
    best_source = None
    best_score = -1e9

    for face_box in mp_faces:
        score = face_inside_head_score(face_box, head_box)
        if score > best_score:
            best_score = score
            best_box = face_box
            best_face_kind = "face"
            best_source = "mp"

    if best_box is None:
        expanded = expand_bbox(head_box, w_img, h_img, scale=1.10)
        if expanded is None:
            expanded = head_box

        x1, y1, x2, y2 = expanded
        roi = frame_bgr[y1:y2, x1:x2]
        if roi.size != 0:
            local_mp_faces = face_detector.detect(roi, detect_width=config.DETECT_WIDTH)
            for fx1, fy1, fx2, fy2 in local_mp_faces:
                gbox = (x1 + fx1, y1 + fy1, x1 + fx2, y1 + fy2)
                if not is_face_candidate_stable(gbox, w_img, h_img):
                    continue

                score = face_inside_head_score(gbox, head_box) + 250.0
                if score > best_score:
                    best_score = score
                    best_box = gbox
                    best_face_kind = "face"
                    best_source = "mp_roi"

    for item in yolo_faces:
        face_box = item["bbox"]
        score = face_inside_head_score(face_box, head_box)
        if score > best_score:
            best_score = score
            best_box = face_box
            best_face_kind = item["kind"]
            best_source = "yolo"

    if best_box is None:
        return None, None, None

    if best_source == "yolo":
        refined = shrink_box(best_box, w_img, h_img, sx=0.90, sy=0.90)
        if refined is not None and is_face_candidate_stable(refined, w_img, h_img):
            best_box = refined

    if best_box is not None and not is_face_candidate_stable(best_box, w_img, h_img):
        return None, None, None

    return best_box, best_face_kind, best_source


def compute_face_metrics(lms, mesh_w, mesh_h, now: float, st: PersonState):
    ear_l = eye_aspect_ratio(lms, LEFT_EYE, mesh_w, mesh_h)
    ear_r = eye_aspect_ratio(lms, RIGHT_EYE, mesh_w, mesh_h)
    ear = (ear_l + ear_r) / 2.0
    is_closed = ear < config.EYES_CLOSED_EAR_T

    st.perclos_events.append((now, is_closed))
    st.perclos_events = [
        (t, c)
        for (t, c) in st.perclos_events
        if t >= now - (config.PERCLOS_WINDOW_S + 2)
    ]

    long_closure = False
    if is_closed and not st.eye_closed:
        st.eye_closed = True
        st.eye_closed_start = now
    elif (not is_closed) and st.eye_closed:
        st.eye_closed = False
        st.eye_closed_start = None

    if st.eye_closed and st.eye_closed_start is not None:
        if (now - st.eye_closed_start) * 1000.0 >= config.DROWSY_CLOSED_MS:
            long_closure = True

    perclos = perclos_from_events(st.perclos_events, now, config.PERCLOS_WINDOW_S)
    fat_pct = fatigue_percent(perclos, long_closure)

    gz = gaze_ratio(lms, mesh_w, mesh_h)
    att_score = attentive_from_gaze(gz, config.LOOK_MIN, config.LOOK_MAX)

    st.last_fatigue_pct = float(fat_pct)
    st.last_attention_pct = float(att_score)
    st.last_seen_ts = now
    st.valid_observations += 1

    return int(fat_pct), float(att_score)